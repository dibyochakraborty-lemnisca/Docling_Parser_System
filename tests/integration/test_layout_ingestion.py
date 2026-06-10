"""Integration: messy xlsx -> layout detector -> observations.

The regression this guards (CRITICAL): a metadata-heavy sheet whose real
header is NOT row 0. Without the detector the parser grabs row 0 as the
header and the file ingests to ZERO observations silently (the praaj case).
With the detector wired, the same file produces observations.

Uses the real pipeline + real ExcelParser + real UnitConverter; fakes only
the repo, file store, and the layout LLM (scripted).
"""

from __future__ import annotations

import uuid
from pathlib import Path

import pytest
from openpyxl import Workbook

from fermdocs.domain.models import DataType, GoldenColumn, GoldenSchema
from fermdocs.file_store.base import StoredFile
from fermdocs.mapping.layout_detector import LayoutDetector
from fermdocs.mapping.mapper import FakeHeaderMapper
from fermdocs.parsing.excel_parser import ExcelParser
from fermdocs.parsing.router import FormatRouter
from fermdocs.pipeline import IngestionPipeline
from fermdocs.units.converter import UnitConverter


class _FakeRepo:
    def __init__(self):
        self.observations = []
        self.residuals = []

    def upsert_experiment(self, *a, **k):
        pass

    def find_or_create_file(self, record):
        return record.file_id, True

    def mark_file_parsed(self, *a, **k):
        pass

    def write_observations(self, obs):
        self.observations.extend(obs)
        return len(obs)

    def write_residual(self, file_id, exp_id, payload, ver):
        self.residuals.append(payload)
        return uuid.uuid4()


class _FakeStore:
    def put(self, src):
        return StoredFile(
            sha256="deadbeef", storage_path=str(src), size_bytes=src.stat().st_size
        )

    def open(self, p):
        return Path(p).read_bytes()


class _ScriptedClient:
    """Returns a layout payload for detect() and a factor payload for the
    cross-run design-factor pass (selected by `kind`)."""

    def __init__(self, payload, factor_payload=None):
        self._payload = payload
        self._factor_payload = factor_payload or {"factors": []}

    def call(self, system, user, *, kind="layout"):
        return self._factor_payload if kind == "factors" else self._payload


def _schema() -> GoldenSchema:
    return GoldenSchema(
        version="test-1.0",
        columns=[
            GoldenColumn(
                name="biomass_g_l", description="biomass", data_type=DataType.FLOAT,
                canonical_unit="g/L", synonyms=["biomass"],
            ),
            GoldenColumn(
                name="temperature_k", description="temp", data_type=DataType.FLOAT,
                canonical_unit="K", synonyms=["temp", "temperature"],
            ),
        ],
    )


def _messy_xlsx(path: Path, sheets=("B471",)) -> Path:
    wb = Workbook()
    wb.remove(wb.active)
    for name in sheets:
        ws = wb.create_sheet(name)
        ws.append(["Praaj Lab Report"])          # row 0: title
        ws.append([f"Batch {name}"])             # row 1: metadata
        ws.append([])                            # row 2: blank
        ws.append(["Time (h)", "Biomass", "Temp"])  # row 3: real header
        ws.append(["0", "0.5", "297"])
        ws.append(["2", "1.2", "298"])
    wb.save(path)
    return path


def _pipeline(detector, repo):
    return IngestionPipeline(
        router=FormatRouter([ExcelParser()]),
        mapper=FakeHeaderMapper(),
        unit_converter=UnitConverter(),
        repository=repo,
        file_store=_FakeStore(),
        schema=_schema(),
        layout_detector=detector,
    )


_GOOD_REGION = {
    "tables": [
        {
            "header_row": 3,
            "data_start_row": 4,
            "data_end_row": 5,
            "run_grouping": "single_run",
            "header_cells": ["Time (h)", "Biomass", "Temp"],
            "confidence": 0.9,
        }
    ]
}


def test_without_detector_messy_sheet_yields_zero_obs(tmp_path):
    # Documents the bug: row-0-as-header assumption -> nothing maps.
    path = _messy_xlsx(tmp_path / "praaj.xlsx")
    repo = _FakeRepo()
    result = _pipeline(detector=None, repo=repo).ingest("exp", [path])
    file_res = result.files[0]
    assert file_res.observations_written == 0
    assert file_res.ingestion_warning == "no_observations_extracted"


def test_with_detector_messy_sheet_ingests(tmp_path):
    # The fix: detector finds the real header row -> observations appear.
    path = _messy_xlsx(tmp_path / "praaj.xlsx")
    repo = _FakeRepo()
    det = LayoutDetector(_ScriptedClient(_GOOD_REGION))
    result = _pipeline(detector=det, repo=repo).ingest("exp", [path])
    file_res = result.files[0]
    assert file_res.observations_written == 4  # biomass + temp, 2 rows each
    assert file_res.ingestion_warning is None
    cols = {o.column_name for o in repo.observations}
    assert cols == {"biomass_g_l", "temperature_k"}


def test_coverage_gate_fires_when_detector_finds_nothing(tmp_path):
    # Detector returns no regions -> fall back to naive table -> 0 obs ->
    # the gate must surface it loudly, not return a clean empty result.
    path = _messy_xlsx(tmp_path / "praaj.xlsx")
    repo = _FakeRepo()
    det = LayoutDetector(_ScriptedClient({"tables": []}))
    result = _pipeline(detector=det, repo=repo).ingest("exp", [path])
    assert result.files[0].observations_written == 0
    assert result.files[0].ingestion_warning == "no_observations_extracted"


def test_design_factors_flow_into_run_conditions(tmp_path):
    # Cross-run design factors (a knob that VARIES across runs, grounded in
    # each sheet's text) become per-run conditions the recommendation reads.
    from openpyxl import Workbook
    path = tmp_path / "campaign.xlsx"
    loadings = {"B471": "15", "B541": "20", "B627": "40"}
    wb = Workbook(); wb.remove(wb.active)
    for name, load in loadings.items():
        ws = wb.create_sheet(name)
        ws.append(["Campaign Report"])
        ws.append([f"Pretreated DBY at {load} g/L"])   # the varying factor
        ws.append([])
        ws.append(["Time (h)", "Biomass", "Temp"])
        ws.append(["0", "0.5", "297"]); ws.append(["2", "1.2", "298"])
    wb.save(path)
    factor_payload = {"factors": [{
        "name": "nutrient_loading_g_l", "kind": "numeric", "unit": "g/L",
        "values": [{"run_id": r, "value": v, "evidence": f"{v} g/L"}
                   for r, v in loadings.items()],
    }]}
    repo = _FakeRepo()
    det = LayoutDetector(_ScriptedClient(_GOOD_REGION, factor_payload=factor_payload))
    _pipeline(detector=det, repo=repo).ingest("exp", [path])
    conds = repo.residuals[0].run_conditions
    assert set(conds) == {"B471", "B541", "B627"}
    # Faithful: raw value kept as written, clean numeric attached.
    assert conds["B471"]["nutrient_loading_g_l"]["value"] == "15"
    assert conds["B471"]["nutrient_loading_g_l"]["numeric"] == 15.0
    assert conds["B627"]["nutrient_loading_g_l"]["numeric"] == 40.0


def test_missing_markers_do_not_become_observations(tmp_path):
    # Lab sheets use "-"/"N/A" for not-measured cells; these must not become
    # observations (they crash a later float()). Biomass col has one "-".
    path = tmp_path / "praaj.xlsx"
    from openpyxl import Workbook
    wb = Workbook(); wb.remove(wb.active)
    ws = wb.create_sheet("B471")
    ws.append(["Praaj Lab Report"]); ws.append(["Batch B471"]); ws.append([])
    ws.append(["Time (h)", "Biomass", "Temp"])
    ws.append(["0", "-", "297"])      # biomass missing
    ws.append(["2", "1.2", "N/A"])    # temp missing
    wb.save(path)
    repo = _FakeRepo()
    det = LayoutDetector(_ScriptedClient(_GOOD_REGION))
    _pipeline(detector=det, repo=repo).ingest("exp", [path])
    vals = {(o.column_name, str(o.value_raw.get("value"))) for o in repo.observations}
    assert ("biomass_g_l", "-") not in vals
    assert all("N/A" not in str(v) and v != "-" for _, v in vals)
    # the real numeric cells still ingested
    assert ("biomass_g_l", "1.2") in vals


def test_multi_sheet_becomes_multi_run(tmp_path):
    path = _messy_xlsx(tmp_path / "praaj.xlsx", sheets=("B471", "B545"))
    repo = _FakeRepo()
    det = LayoutDetector(_ScriptedClient(_GOOD_REGION))
    _pipeline(detector=det, repo=repo).ingest("exp", [path])
    run_ids = {o.source_locator.get("run_id") for o in repo.observations}
    assert run_ids == {"B471", "B545"}

#!/usr/bin/env python3
"""Convert a Lemnisca lab-report spreadsheet into canonical time-series observations.

The lab reports (e.g. `B471_PHS_CaOH2.csv`) are human documents: a metadata-heavy
header (feedstock analysis, reactor design, media prep) followed by ONE embedded
analysis table (section "D") whose rows are sampling timepoints across several
process stages (Seed -> CV1 -> CV2 -> PF -> MF). Characterize/diagnose ingest
nothing from this shape: the column names aren't canonical, units are %w/w, and
the `Time (Hours)` column resets per stage so there's no single monotonic clock.
Result: 0 trajectories, 0 findings, an empty fault-finding run.

This converter extracts ONE stage (default the main fermentation, `MF`) as a
single run on a monotonic clock and emits the canonical long-format observations
schema the pipeline expects:

    run_id, variable, time_h, value, unit

Conversions / decisions (all explicit, all documented):
  * %w/w -> g/L : value_g_per_L = pct_w_w * 10 * density. Density defaults to
    1.06 (broth density on this sheet; sanity-checked: 9.62 %w/w LA -> 102.0 g/L,
    matching the sheet's own 102.02 g/L conclusion).
  * Substrate = `Total (%w/w)` residual sugar. Product = `Lactic Acid(%w/w)`.
  * Volume `Actual Broth Volume (ml)` -> L.
  * pH passes through.
  * `Volume Corrected OD @ 600nm` and `DO %` are kept as run-only proxy variables
    (od_600nm, dissolved_o2_pct). OD is NOT dry-cell-weight g/L; it is emitted
    honestly under its own name rather than faked into biomass_g_l.

Stage selection: rows whose `Sample Name` starts with the stage prefix. The seed
train (Seed/CV/PF) is deliberately dropped — it is not the fermentation under
study and its time axis is separate.

Usage:
    python scripts/convert_lab_report.py INPUT.csv [--stage MF] [--run-id ID]
        [--density 1.06] [--out observations.csv]
"""
from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path


def _norm(s: str) -> str:
    """Normalize a header for tolerant matching: lowercase, alphanumerics only."""
    return re.sub(r"[^a-z0-9]", "", str(s).lower())


# output variable -> (normalized source header, unit, transform)
# transform takes the raw float and returns the canonical value.
def _spec(density: float):
    pct_to_gL = lambda v: v * 10.0 * density  # noqa: E731
    return {
        "substrate_g_l": (_norm("Total (%w/w)"), "g/L", pct_to_gL),
        "product_g_l": (_norm("Lactic Acid(%w/w)"), "g/L", pct_to_gL),
        "volume_l": (_norm("Actual Broth Volume (ml)"), "L", lambda v: v / 1000.0),
        "ph": (_norm("pH"), "pH", lambda v: v),
        "od_600nm": (_norm("Volume Corrected OD @ 600nm"), "OD600", lambda v: v),
        "dissolved_o2_pct": (_norm("DO %"), "%", lambda v: v),
    }


_TIME_KEY = _norm("Time (Hours)")
_SAMPLE_KEY = _norm("Sample Name")


def _to_float(cell: str):
    """Parse a numeric cell, tolerating blanks, '-', and stray spaces. None if not numeric."""
    s = str(cell).strip()
    if s in ("", "-", "ND", "NA", "N/A"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _read_rows(path: Path) -> list[list[str]]:
    with path.open(newline="", encoding="utf-8-sig") as fh:
        return [row for row in csv.reader(fh)]


def _iter_sheets(path: Path):
    """Yield (sheet_name, rows) for each sheet. A .csv is a single unnamed sheet
    (sheet_name=None); an .xlsx yields one (name, rows) per worksheet so a 15-sheet
    workbook becomes 15 runs."""
    if path.suffix.lower() in (".xlsx", ".xls"):
        import openpyxl  # lazy: only needed for workbooks
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for name in wb.sheetnames:
            ws = wb[name]
            rows = [["" if c is None else c for c in row]
                    for row in ws.iter_rows(values_only=True)]
            yield name, rows
    else:
        yield None, _read_rows(path)


def _find_header(rows: list[list[str]]) -> tuple[int, dict[str, int]]:
    """Locate the analysis table: the row carrying both Sample Name and Time (Hours).
    Returns (row_index, {normalized_header: column_index})."""
    for i, row in enumerate(rows):
        norm_cells = {_norm(c): j for j, c in enumerate(row) if str(c).strip()}
        if _SAMPLE_KEY in norm_cells and _TIME_KEY in norm_cells:
            return i, norm_cells
    raise SystemExit("ERROR: could not find the analysis table "
                     "(no row with both 'Sample Name' and 'Time (Hours)').")


def _rows_to_records(rows: list[list], *, stage: str, run_id: str, density: float):
    """Extract one stage of one table into canonical observation records."""
    hdr_idx, cols = _find_header(rows)
    specs = _spec(density)
    stage_key = _norm(stage)

    records: list[dict] = []
    n_points = 0
    for row in rows[hdr_idx + 1:]:
        if len(row) <= cols[_SAMPLE_KEY]:
            continue
        sample = str(row[cols[_SAMPLE_KEY]]).strip()
        if not sample:
            continue  # blank line -> end of table region (we still scan on, harmless)
        if not _norm(sample).startswith(stage_key):
            continue
        t = _to_float(row[cols[_TIME_KEY]])
        if t is None:
            continue
        n_points += 1
        for var, (src_key, unit, fn) in specs.items():
            if src_key not in cols:
                continue
            raw = _to_float(row[cols[src_key]])
            if raw is None:
                continue
            records.append({
                "run_id": run_id, "variable": var,
                "time_h": t, "value": round(fn(raw), 4), "unit": unit,
            })
    return records, n_points


def convert(path: Path, *, stage: str, run_id: str | None, density: float):
    """Convert every sheet (one run per sheet for an xlsx; one run for a csv).
    Returns (records, per_run_points) where per_run_points maps run_id -> n_timepoints."""
    records: list[dict] = []
    per_run: dict[str, int] = {}
    for sheet_name, rows in _iter_sheets(path):
        rid = re.sub(r"[^A-Za-z0-9_]+", "_", (sheet_name or run_id or path.stem)).strip("_")
        try:
            recs, n_points = _rows_to_records(rows, stage=stage, run_id=rid, density=density)
        except SystemExit as exc:  # no analysis table on this sheet — skip, don't abort the workbook
            print(f"  [skip] {rid}: {exc}", file=sys.stderr)
            continue
        if recs:
            records.extend(recs)
            per_run[rid] = n_points
    return records, per_run


def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("input", type=Path, help="lab-report CSV")
    ap.add_argument("--stage", default="MF",
                    help="Sample-Name prefix of the stage to extract (default: MF = main fermentation)")
    ap.add_argument("--run-id", default=None,
                    help="run id for the emitted rows (default: derived from filename)")
    ap.add_argument("--density", type=float, default=1.06,
                    help="broth density for %%w/w -> g/L conversion (default 1.06)")
    ap.add_argument("--out", type=Path, default=None,
                    help="output CSV (default: <input_stem>.observations.csv)")
    ap.add_argument("--format", choices=["long", "wide"], default="long",
                    help="long = canonical observations (run_id,variable,time_h,value,unit); "
                         "wide = one row per (run,timepoint) with IndPenSim-style headers "
                         "(Batch_ref,Time (h),S,P,V,pH) that the upload pipeline ingests directly")
    args = ap.parse_args(argv)

    if not args.input.exists():
        raise SystemExit(f"ERROR: input not found: {args.input}")
    out = args.out or args.input.with_suffix("").with_name(args.input.stem + ".observations.csv")

    records, per_run = convert(args.input, stage=args.stage, run_id=args.run_id, density=args.density)
    if not records:
        raise SystemExit(f"ERROR: no '{args.stage}' rows produced any observations "
                         f"(check the stage prefix and the analysis table).")

    if args.format == "wide":
        # Wide table the existing upload pipeline ingests directly: one row per
        # (run, timepoint). Headers are the EXACT golden-schema raw_headers so the
        # mapper resolves them unambiguously (a bare "P"/"Batch_ref" is ambiguous
        # to the LLM mapper and gets dropped). Batch -> experiment_id/run grouping.
        wide_map = {
            "substrate_g_l": "Substrate (S)",
            "product_g_l": "Product (g/L)",
            "volume_l": "Volume",
            "ph": "pH",
        }
        cells: dict[tuple[str, float], dict[str, float]] = {}
        for r in records:
            col = wide_map.get(r["variable"])
            if col is None:
                continue
            cells.setdefault((r["run_id"], float(r["time_h"])), {})[col] = r["value"]
        fields = ["Batch", "Time (h)", *wide_map.values()]
        with out.open("w", newline="", encoding="utf-8") as fh:
            w = csv.DictWriter(fh, fieldnames=fields)
            w.writeheader()
            for (rid, t), vals in sorted(cells.items()):
                w.writerow({"Batch": rid, "Time (h)": t, **vals})
        n_rows = len(cells)
    else:
        with out.open("w", newline="", encoding="utf-8") as fh:
            w = csv.DictWriter(fh, fieldnames=["run_id", "variable", "time_h", "value", "unit"])
            w.writeheader()
            w.writerows(records)
        n_rows = len(records)

    print(f"stage={args.stage}  format={args.format}  runs={len(per_run)}  rows={n_rows}")
    for rid, n_points in sorted(per_run.items()):
        nvars = len({r["variable"] for r in records if r["run_id"] == rid})
        print(f"  {rid:24s} timepoints={n_points:2d}  variables={nvars}")
    print(f"wrote {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
